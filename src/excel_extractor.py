# excel_extractor.py
from openpyxl import load_workbook
from typing import Iterator
from extract_base import Extractor, Record

class Extractor(Extractor):
    def __init__(self, path: str, header_row_index: int = 1):
        """
        path: path to the Excel file
        header_row_index: which row has the headers (1-based, usually 1)
        """
        self.path = path
        self.header_row_index = header_row_index
        self.columns = {}  # will be filled automatically

        # Auto-map columns from header row
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb.active  # always first sheet
        self.map_columns(ws)
        wb.close()

    def map_columns(self, ws) -> None:
        """Build self.columns = {title: index} from the header row."""
        self.columns.clear()
        for i, title in enumerate(ws[self.header_row_index]):
            if title.value is None:
                continue
            name = str(title.value).strip()
            if name:
                self.columns.setdefault(name, i)

    def records(self) -> Iterator[Record]:
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb.active  # always first sheet

            # Resolve indices once (faster per-row)
            idx_sample_id    = self.columns.get("Sample ID")
            idx_sample_type  = self.columns.get("Sample Type")
            idx_mean         = self.columns.get("Mean (per analysis type)")
            idx_ppm          = self.columns.get("PPM")
            idx_adjusted_abs = self.columns.get("Adjusted ABS")

            for row in ws.iter_rows(min_row=self.header_row_index + 1, values_only=True):
                def get(idx: int | None):
                    return None if idx is None or idx >= len(row) else row[idx]

                sample_id    = self.to_str(get(idx_sample_id))
                sample_type  = self.to_str(get(idx_sample_type))
                mean         = self.to_float(get(idx_mean))
                ppm          = self.to_float(get(idx_ppm))
                adjusted_abs = self.to_float(get(idx_adjusted_abs))

                if not sample_id and mean is None and ppm is None and adjusted_abs is None:
                    continue

                yield Record(
                    sample_id=sample_id,
                    sample_type=sample_type,
                    mean=mean,
                    ppm=ppm,
                    adjusted_abs=adjusted_abs
                )
        finally:
            wb.close()

    @staticmethod
    def to_float(value) -> float | None:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def to_str(value) -> str:
        return "" if value is None else str(value).strip()



'''
# Example use
ex = Extractor("test_file.xlsx", header_row_index = 1)
for rec in ex.records():
    print(rec.sample_id, rec.sample_type, rec.mean, rec.ppm, rec.adjusted_abs)

'''