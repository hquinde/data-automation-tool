import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from excel_extract import Extract
from excel_transform import Transform


class Load:
    def __init__(self, transformer):
        self.transformer = transformer
        self.output_path = "results.xlsx"
        self.molecular_weight = 12.01057


    @staticmethod
    def is_out_of_bounds(value, check_type):
        val = float(value)
        
        if check_type == 'QC_R':
            return val < 90 or val > 110
        elif check_type == 'MDL_R':
            return val < 45 or val > 145
        elif check_type == 'RPD':
            return val > 10
        else:
            return True


    def sample_groups(self):
        cleaned = self.transformer.clean_data()
        
        df = cleaned.copy()
        df["Sample ID"] = df["Sample ID"].astype("string").str.strip()

        qc_pattern = r"(?i)^(MDL|QCS|QCB|CCV\d+|CCB\d+|MQ\s*(?:Auto)?Rinse)$"
        samples_only = df[~df["Sample ID"].str.match(qc_pattern, na=False)]

        ordered_ids = self.get_unique_ordered_ids(samples_only)
        groups = self.build_sample_groups(samples_only, ordered_ids)

        return samples_only, groups


    def get_unique_ordered_ids(self, df):
        ordered_ids = []
        seen = set()
        for sid in df["Sample ID"]:
            if sid not in seen:
                seen.add(sid)
                ordered_ids.append(sid)
        return ordered_ids


    def build_sample_groups(self, df, ordered_ids):
        groups = []
        for sample_id in ordered_ids:
            group_df = df[df["Sample ID"] == sample_id]
            groups.append((sample_id, group_df))
        return groups


    def format_qc(self):
        df = self.transformer.df
        samples = df[df.get("Sample Type") == "Samples"]
        samples["Sample ID"] = samples["Sample ID"].astype("string").str.strip()

        qc_mask = samples["Sample ID"].str.match(r"(?i)^(MDL|QCS|CCV\d+)$", na=False)
        qcb_mask = samples["Sample ID"].str.match(r"(?i)^(QCB|CCB\d+)$", na=False)

        qc_samples = samples[qc_mask]
        qcb_samples = samples[qcb_mask]

        columns = ["Sample ID", "PPM C", "Mean ppm C", "%R", "%RPD"]
        
        records = []
        records.extend(self.build_qc_records(qc_samples))
        records.append({col: None for col in columns})
        records.extend(self.build_qcb_records(qcb_samples))
        records.append(self.build_qcb_average(qcb_samples))

        return pd.DataFrame(records, columns=columns)


    def build_qc_records(self, qc_samples):
        qc_targets = {
            "MDL": 0.2,
            "QCS": 18.0,
        }
        
        records = []
        for sample_id in qc_samples["Sample ID"].unique():
            group_df = qc_samples[qc_samples["Sample ID"] == sample_id]

            group_records = []
            for _, row in group_df.iterrows():
                group_records.append(
                    {
                        "Sample ID": row["Sample ID"],
                        "PPM C": row.get("PPM"),
                        "Mean ppm C": None,
                        "%R": None,
                        "%RPD": None,
                    }
                )

            mean_ppm = self.transformer.calculate_mean_ppm(group_df)
            
            sample_id_upper = str(sample_id).upper()
            if sample_id_upper.startswith("CCV"):
                target = 10.0
            else:
                target = qc_targets.get(sample_id_upper)

            percent_r = self.transformer.calculate_percent_R(group_df, target_override=target)
            rpd = self.transformer.calculate_rpd(group_df, mean_ppm)

            last_record = group_records[-1]
            last_record["Mean ppm C"] = mean_ppm
            last_record["%R"] = percent_r
            last_record["%RPD"] = rpd

            records.extend(group_records)
        
        return records


    def build_qcb_records(self, qcb_samples):
        records = []
        for _, row in qcb_samples.iterrows():
            records.append(
                {
                    "Sample ID": row["Sample ID"],
                    "PPM C": row.get("PPM"),
                    "Mean ppm C": None,
                    "%R": None,
                    "%RPD": None,
                }
            )
        return records


    def build_qcb_average(self, qcb_samples):
        average_ppm = self.transformer.calculate_mean_ppm(qcb_samples)
        return {
            "Sample ID": "Average",
            "PPM C": average_ppm,
            "Mean ppm C": None,
            "%R": None,
            "%RPD": None,
        }


    def format_samples(self):
        samples_only, groups = self.sample_groups()
        
        columns = ["Sample ID", "PPM C", "Mean ppm C", "%RPD", "umol/L C"]
        records = []

        for sample_id, group_df in groups:
            group_records = self.build_sample_group_records(group_df)
            self.add_summary_to_last_record(group_df, group_records)
            records.extend(group_records)

        return pd.DataFrame(records, columns=columns)


    def build_sample_group_records(self, group_df):
        group_records = []
        for _, row in group_df.iterrows():
            group_records.append(
                {
                    "Sample ID": row["Sample ID"],
                    "PPM C": row.get("PPM"),
                    "Mean ppm C": None,
                    "%RPD": None,
                    "umol/L C": None,
                }
            )
        return group_records


    def add_summary_to_last_record(self, group_df, group_records):
        mean_ppm = self.transformer.calculate_mean_ppm(group_df)
        rpd = self.transformer.calculate_rpd(group_df, mean_ppm)
        mean_umol = self.transformer.convert_to_umol_per_L(mean_ppm, self.molecular_weight)

        last_record = group_records[-1]
        last_record["Mean ppm C"] = mean_ppm
        last_record["%RPD"] = rpd
        last_record["umol/L C"] = mean_umol


    def format_reported_results(self):
        _, groups = self.sample_groups()
        
        records = []
        for sample_id, group_df in groups:
            mean_ppm = self.transformer.calculate_mean_ppm(group_df)
            umol = self.transformer.convert_to_umol_per_L(mean_ppm, self.molecular_weight)

            records.append(
                {
                    "Sample ID": sample_id,
                    "umol/L C": umol,
                }
            )

        return pd.DataFrame(records, columns=["Sample ID", "umol/L C"])


    def export_all(self):
        self.write_sheets()
        self.apply_formatting()
        print(f"Export finished: {self.output_path}")


    def write_sheets(self):
        sheets = {
            "Raw Data": self.transformer.df,
            "QC": self.format_qc(),
            "Samples": self.format_samples(),
            "Reported Results": self.format_reported_results(),
        }

        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Wrote {len(df)} rows to sheet '{sheet_name}'")


    def apply_formatting(self):
        wb = load_workbook(self.output_path)
        red_font = Font(color='FF0000', bold=True)
        
        self.format_qc_sheet(wb, red_font)
        self.format_samples_sheet(wb, red_font)
        
        wb.save(self.output_path)
        print(f"Applied bounds checking and formatting")


    def format_qc_sheet(self, wb, red_font):
        ws = wb['QC']
        for row_idx in range(2, ws.max_row + 1):
            sample_id_cell = ws.cell(row=row_idx, column=1)
            sample_id = sample_id_cell.value
            
            r_cell = ws.cell(row=row_idx, column=4)
            if r_cell.value is not None:
                if sample_id and 'MDL' in str(sample_id).upper():
                    check_type = 'MDL_R'
                else:
                    check_type = 'QC_R'
                
                if self.is_out_of_bounds(r_cell.value, check_type):
                    r_cell.font = red_font


    def format_samples_sheet(self, wb, red_font):
        ws = wb['Samples']
        for row_idx in range(2, ws.max_row + 1):
            rpd_cell = ws.cell(row=row_idx, column=4)
            if rpd_cell.value is not None:
                if self.is_out_of_bounds(rpd_cell.value, 'RPD'):
                    rpd_cell.font = red_font


if __name__ == "__main__":
    extractor = Extract("POSTER.xlsx")
    raw_data = extractor.extract_data()

    if raw_data is not None:
        transformer = Transform(raw_data)
        loader = Load(transformer)
        loader.export_all()
    else:
        print("Failed to load data. Check file path and format.")